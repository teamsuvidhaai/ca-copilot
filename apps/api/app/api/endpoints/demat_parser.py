"""
Zerodha Demat Direct Parser
────────────────────────────
Parses Zerodha-format Excel exports directly without AI.
Handles: Holdings, Tax P&L (tradewise exits), Tradebook.
Produces the same structured JSON as the GPT-4o pipeline.

Zero API cost. Instant. 100% reliable.
"""

import io
import logging
from datetime import datetime, date
from decimal import Decimal
from typing import Optional

import openpyxl

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════
# BROKER DETECTION
# ═══════════════════════════════════════════════════════

def detect_broker(wb: openpyxl.Workbook) -> str:
    """Detect broker from workbook content."""
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for r in range(1, min(15, ws.max_row + 1)):
            for c in range(1, min(5, ws.max_column + 1)):
                val = str(ws.cell(r, c).value or "").lower()
                if "zerodha" in val:
                    return "zerodha"
                if "groww" in val:
                    return "groww"
                if "angel" in val:
                    return "angel"
    # Check for Zerodha patterns (Client ID format, typical sheet names)
    if any(s in ["Equity", "Combined"] for s in wb.sheetnames):
        return "zerodha"
    return "unknown"


# ═══════════════════════════════════════════════════════
# HEADER FINDER — locates the data header row
# ═══════════════════════════════════════════════════════

def find_header_row(ws, target_headers: list[str], max_search: int = 30) -> Optional[int]:
    """Find the row number containing the header columns."""
    targets_lower = [h.lower() for h in target_headers]
    for r in range(1, min(max_search, ws.max_row + 1)):
        row_vals = [str(ws.cell(r, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]
        matches = sum(1 for t in targets_lower if t in row_vals)
        if matches >= len(targets_lower) * 0.6:  # 60% header match
            return r
    return None


def get_col_map(ws, header_row: int) -> dict:
    """Build {lowercase_header: column_index} map."""
    cols = {}
    for c in range(1, ws.max_column + 1):
        val = str(ws.cell(header_row, c).value or "").strip()
        if val:
            cols[val.lower()] = c
    return cols


def cell_val(ws, row: int, col: int, as_type=str):
    """Safely read a cell value with type conversion."""
    v = ws.cell(row, col).value
    if v is None:
        return None
    if as_type == float:
        try:
            return float(str(v).replace(",", ""))
        except (ValueError, TypeError):
            return None
    if as_type == str:
        return str(v).strip()
    return v


# ═══════════════════════════════════════════════════════
# HOLDINGS PARSER
# ═══════════════════════════════════════════════════════

def parse_holdings(file_bytes: bytes, filename: str) -> dict:
    """Parse Zerodha holdings Excel → structured JSON."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    broker = detect_broker(wb)
    result = {
        "broker": broker,
        "client_id": None,
        "statement_date": None,
        "holdings": [],
        "summary": {},
    }

    # Extract client ID
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for r in range(1, 15):
            if str(ws.cell(r, 2).value or "").lower() == "client id":
                result["client_id"] = str(ws.cell(r, 3).value or "")
                break

    # Parse Equity sheet
    if "Equity" in wb.sheetnames:
        ws = wb["Equity"]
        header_row = find_header_row(ws, ["Symbol", "ISIN", "Quantity Available", "Average Price"])
        if header_row:
            cols = get_col_map(ws, header_row)
            for r in range(header_row + 1, ws.max_row + 1):
                symbol = cell_val(ws, r, cols.get("symbol", 2))
                if not symbol:
                    continue
                holding = {
                    "scrip_name": symbol,
                    "isin": cell_val(ws, r, cols.get("isin", 3)),
                    "sector": cell_val(ws, r, cols.get("sector", 4)),
                    "quantity": cell_val(ws, r, cols.get("quantity available", 5), float),
                    "quantity_long_term": cell_val(ws, r, cols.get("quantity long term", 7), float),
                    "avg_cost": cell_val(ws, r, cols.get("average price", 10), float),
                    "market_price": cell_val(ws, r, cols.get("previous closing price", 11), float),
                    "unrealised_gain": cell_val(ws, r, cols.get("unrealized p&l", 12), float),
                    "unrealised_gain_pct": cell_val(ws, r, cols.get("unrealized p&l pct.", 13), float),
                }
                if holding["quantity"] and holding.get("avg_cost"):
                    holding["cost_value"] = round(holding["quantity"] * holding["avg_cost"], 2)
                if holding["quantity"] and holding.get("market_price"):
                    holding["market_value"] = round(holding["quantity"] * holding["market_price"], 2)
                result["holdings"].append(holding)

        # Extract summary
        for r in range(1, 20):
            key = str(ws.cell(r, 2).value or "").lower()
            val = ws.cell(r, 3).value
            if "invested value" in key and val:
                result["summary"]["invested_value"] = float(str(val).replace(",", ""))
            elif "present value" in key and val:
                result["summary"]["present_value"] = float(str(val).replace(",", ""))
            elif "unrealized p&l" == key.strip() and val:
                result["summary"]["unrealised_pnl"] = float(str(val).replace(",", ""))

    wb.close()
    logger.info(f"✅ Holdings: {len(result['holdings'])} securities parsed (broker={broker})")
    return result


# ═══════════════════════════════════════════════════════
# TAX P&L PARSER
# ═══════════════════════════════════════════════════════

def parse_taxpnl(file_bytes: bytes, filename: str) -> dict:
    """Parse Zerodha Tax P&L Excel → structured JSON with tradewise exits."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    result = {
        "client_id": None,
        "client_name": None,
        "pan": None,
        "period_start": None,
        "period_end": None,
        "transactions": [],
        "dividends": [],
        "capital_gains_summary": {
            "intraday_profit": None,
            "short_term_profit": None,
            "long_term_profit": None,
            "total_profit": None,
        },
    }

    # Extract metadata from first sheet
    ws0 = wb[wb.sheetnames[0]]
    for r in range(1, 15):
        key = str(ws0.cell(r, 2).value or "").lower()
        val = str(ws0.cell(r, 3).value or "")
        if "client id" in key:
            result["client_id"] = val
        elif "client name" in key:
            result["client_name"] = val
        elif "pan" in key:
            result["pan"] = val

    # Parse "Tradewise Exits" sheet
    tradewise_sheet = None
    for s in wb.sheetnames:
        if "tradewise" in s.lower():
            tradewise_sheet = s
            break

    if tradewise_sheet:
        ws = wb[tradewise_sheet]
        header_row = find_header_row(ws, ["Symbol", "ISIN", "Entry Date", "Exit Date", "Buy Value", "Sell Value"])
        if header_row:
            cols = get_col_map(ws, header_row)
            current_section = ""
            for r in range(header_row + 1, ws.max_row + 1):
                # Detect section headers (Equity - Intraday, Equity - Short Term, etc.)
                first_col = str(ws.cell(r, 2).value or "").strip()
                if first_col and not ws.cell(r, 3).value and not ws.cell(r, 5).value:
                    # This is a section header like "Equity - Intraday"
                    current_section = first_col
                    continue

                symbol = cell_val(ws, r, cols.get("symbol", 2))
                if not symbol:
                    continue

                entry_date = cell_val(ws, r, cols.get("entry date", 4))
                exit_date = cell_val(ws, r, cols.get("exit date", 5))
                quantity = cell_val(ws, r, cols.get("quantity", 6), float)
                buy_value = cell_val(ws, r, cols.get("buy value", 7), float)
                sell_value = cell_val(ws, r, cols.get("sell value", 8), float)
                profit = cell_val(ws, r, cols.get("profit", 9), float)
                holding_days = cell_val(ws, r, cols.get("period of holding", 10), float)

                # Determine holding period category
                if holding_days is not None:
                    if holding_days == 0:
                        holding_period = "Intraday"
                    elif holding_days <= 365:
                        holding_period = "Short Term"
                    else:
                        holding_period = "Long Term"
                else:
                    holding_period = None

                txn = {
                    "scrip_name": symbol,
                    "isin": cell_val(ws, r, cols.get("isin", 3)),
                    "entry_date": str(entry_date) if entry_date else None,
                    "exit_date": str(exit_date) if exit_date else None,
                    "quantity": quantity,
                    "buy_value": buy_value,
                    "sell_value": sell_value,
                    "profit": profit,
                    "holding_days": int(holding_days) if holding_days is not None else None,
                    "holding_period": holding_period,
                    "section": current_section,
                    "brokerage": cell_val(ws, r, cols.get("brokerage", 14), float),
                    "stt": cell_val(ws, r, cols.get("stt", 16), float) if cols.get("stt") else None,
                }
                result["transactions"].append(txn)

    # Parse summary from "Equity and Non Equity" sheet
    eq_sheet = None
    for s in wb.sheetnames:
        if "equity and non" in s.lower():
            eq_sheet = s
            break

    if eq_sheet:
        ws = wb[eq_sheet]
        for r in range(1, ws.max_row + 1):
            key = str(ws.cell(r, 2).value or "").strip().lower()
            val = ws.cell(r, 3).value
            if val is None:
                continue
            try:
                fval = float(str(val).replace(",", ""))
            except (ValueError, TypeError):
                continue
            if "intraday" in key and "profit" in key:
                result["capital_gains_summary"]["intraday_profit"] = fval
            elif "short term profit" in key:
                result["capital_gains_summary"]["short_term_profit"] = fval
            elif "long term profit" in key:
                result["capital_gains_summary"]["long_term_profit"] = fval

    cg = result["capital_gains_summary"]
    cg["total_profit"] = sum(v or 0 for v in [cg.get("intraday_profit"), cg.get("short_term_profit"), cg.get("long_term_profit")])

    # Parse dividends sheet
    for s in wb.sheetnames:
        if "dividend" in s.lower():
            ws = wb[s]
            header_row = find_header_row(ws, ["Symbol", "Dividend"])
            if header_row:
                cols = get_col_map(ws, header_row)
                for r in range(header_row + 1, ws.max_row + 1):
                    symbol = cell_val(ws, r, cols.get("symbol", 2))
                    if not symbol:
                        continue
                    result["dividends"].append({
                        "scrip_name": symbol,
                        "isin": cell_val(ws, r, cols.get("isin", 3)) if cols.get("isin") else None,
                        "date": cell_val(ws, r, cols.get("ex date", 4) or cols.get("date", 4)),
                        "amount": cell_val(ws, r, cols.get("dividend", 5) or cols.get("amount", 5), float),
                    })
            break

    wb.close()
    logger.info(f"✅ Tax P&L: {len(result['transactions'])} exits, {len(result['dividends'])} dividends, CG={result['capital_gains_summary']}")
    return result


# ═══════════════════════════════════════════════════════
# TRADEBOOK PARSER
# ═══════════════════════════════════════════════════════

def parse_tradebook(file_bytes: bytes, filename: str) -> dict:
    """Parse Zerodha tradebook Excel → structured JSON."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    result = {
        "client_id": None,
        "transactions": [],
    }

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # Find client ID
        for r in range(1, 15):
            if str(ws.cell(r, 2).value or "").lower() == "client id":
                result["client_id"] = str(ws.cell(r, 3).value or "")
                break

        # Zerodha tradebook headers:
        # Symbol, ISIN, Trade Date, Exchange, Segment, Series, Trade Type, Auction, Quantity, Price, Trade ID, Order ID, Order Execution Time
        header_row = find_header_row(ws, ["Symbol", "ISIN", "Trade Date", "Trade Type", "Quantity", "Price"])
        if not header_row:
            continue

        cols = get_col_map(ws, header_row)
        for r in range(header_row + 1, ws.max_row + 1):
            symbol = cell_val(ws, r, cols.get("symbol", 2))
            if not symbol:
                continue

            qty = cell_val(ws, r, cols.get("quantity", 10), float)
            price = cell_val(ws, r, cols.get("price", 11), float)

            txn = {
                "scrip_name": symbol,
                "isin": cell_val(ws, r, cols.get("isin", 3)),
                "date": cell_val(ws, r, cols.get("trade date", 4)),
                "exchange": cell_val(ws, r, cols.get("exchange", 5)),
                "segment": cell_val(ws, r, cols.get("segment", 6)),
                "type": (cell_val(ws, r, cols.get("trade type", 8)) or "").capitalize(),  # Buy/Sell
                "quantity": qty,
                "price": price,
                "amount": round(qty * price, 2) if qty and price else None,
                "trade_id": cell_val(ws, r, cols.get("trade id", 12)),
                "order_time": cell_val(ws, r, cols.get("order execution time", 14)),
            }
            result["transactions"].append(txn)

    wb.close()
    logger.info(f"✅ Tradebook: {len(result['transactions'])} trades parsed")
    return result


# ═══════════════════════════════════════════════════════
# JOURNAL ENTRY GENERATOR (Rule-Based, No AI)
# ═══════════════════════════════════════════════════════

def generate_journal_entries_rules(structured_data: dict, instrument_subtype: str) -> list:
    """Generate journal entries using accounting rules — no AI needed."""
    entries = []

    if instrument_subtype == "demat_taxpnl":
        # Tax P&L → Capital gain/loss entries per exit
        for txn in structured_data.get("transactions", []):
            profit = txn.get("profit")
            if profit is None or profit == 0:
                continue

            symbol = txn.get("scrip_name", "Unknown")
            exit_date = txn.get("exit_date", txn.get("date"))
            buy_val = txn.get("buy_value", 0) or 0
            sell_val = txn.get("sell_value", 0) or 0
            holding_period = txn.get("holding_period", "Short Term")

            if holding_period == "Intraday":
                gain_ledger = "Speculative Income" if profit > 0 else "Speculative Loss"
            elif holding_period == "Long Term":
                gain_ledger = "LTCG on Shares" if profit > 0 else "LTCL on Shares"
            else:
                gain_ledger = "STCG on Shares" if profit > 0 else "STCL on Shares"

            ledger_entries = [
                {"ledger_name": "Bank Account", "amount": round(abs(sell_val), 2), "side": "Dr"},
                {"ledger_name": "Investment in Equity Shares", "amount": round(abs(buy_val), 2), "side": "Cr"},
            ]

            if profit > 0:
                ledger_entries.append({"ledger_name": gain_ledger, "amount": round(abs(profit), 2), "side": "Cr"})
            else:
                ledger_entries.append({"ledger_name": gain_ledger, "amount": round(abs(profit), 2), "side": "Dr"})

            entries.append({
                "date": str(exit_date) if exit_date else None,
                "voucher_type": "Sales",
                "narration": f"Sale of {txn.get('quantity', 0):.0f} {symbol} — {holding_period} {gain_ledger}",
                "ledger_entries": ledger_entries,
            })

        # Dividend entries
        for div in structured_data.get("dividends", []):
            amt = div.get("amount")
            if not amt or amt == 0:
                continue
            entries.append({
                "date": str(div.get("date")) if div.get("date") else None,
                "voucher_type": "Receipt",
                "narration": f"Dividend from {div.get('scrip_name', 'Unknown')}",
                "ledger_entries": [
                    {"ledger_name": "Bank Account", "amount": round(abs(amt), 2), "side": "Dr"},
                    {"ledger_name": "Dividend Income", "amount": round(abs(amt), 2), "side": "Cr"},
                ],
            })

    elif instrument_subtype == "demat_tradebook":
        # Tradebook → Buy/Sell vouchers
        for txn in structured_data.get("transactions", []):
            symbol = txn.get("scrip_name", "Unknown")
            trade_type = (txn.get("type") or "").lower()
            amount = txn.get("amount") or 0
            if amount == 0:
                continue

            if trade_type == "buy":
                entries.append({
                    "date": str(txn.get("date")) if txn.get("date") else None,
                    "voucher_type": "Purchase",
                    "narration": f"Purchase of {txn.get('quantity', 0):.0f} {symbol} @ ₹{txn.get('price', 0):.2f}",
                    "ledger_entries": [
                        {"ledger_name": "Investment in Equity Shares", "amount": round(amount, 2), "side": "Dr"},
                        {"ledger_name": "Bank Account", "amount": round(amount, 2), "side": "Cr"},
                    ],
                })
            elif trade_type == "sell":
                entries.append({
                    "date": str(txn.get("date")) if txn.get("date") else None,
                    "voucher_type": "Sales",
                    "narration": f"Sale of {txn.get('quantity', 0):.0f} {symbol} @ ₹{txn.get('price', 0):.2f}",
                    "ledger_entries": [
                        {"ledger_name": "Bank Account", "amount": round(amount, 2), "side": "Dr"},
                        {"ledger_name": "Investment in Equity Shares", "amount": round(amount, 2), "side": "Cr"},
                    ],
                })

    elif instrument_subtype == "demat_holdings":
        # Holdings → Opening balance entries for each security
        for h in structured_data.get("holdings", []):
            symbol = h.get("scrip_name", "Unknown")
            qty = h.get("quantity") or 0
            avg_cost = h.get("avg_cost") or 0
            cost_value = h.get("cost_value") or (qty * avg_cost if qty and avg_cost else 0)
            market_value = h.get("market_value") or 0

            if cost_value <= 0:
                continue

            # Opening balance entry: Dr Investment / Cr Opening Balance
            entries.append({
                "date": None,  # Opening balance — no specific date
                "voucher_type": "Journal",
                "narration": f"Opening balance — {qty:.0f} shares of {symbol} @ ₹{avg_cost:.2f} (Sector: {h.get('sector', 'N/A')})",
                "ledger_entries": [
                    {"ledger_name": f"Investment in Equity - {symbol}", "amount": round(cost_value, 2), "side": "Dr"},
                    {"ledger_name": "Capital / Opening Balance", "amount": round(cost_value, 2), "side": "Cr"},
                ],
            })

    return entries


# ═══════════════════════════════════════════════════════
# MAIN ENTRY POINT
# ═══════════════════════════════════════════════════════

def parse_demat_excel(file_bytes: bytes, filename: str, instrument_type: str) -> tuple[dict, list]:
    """
    Parse a Demat Excel file and generate journal entries.
    Returns: (structured_data, journal_entries)
    No AI/API calls — pure local processing.
    """
    if instrument_type == "demat_holdings":
        structured = parse_holdings(file_bytes, filename)
    elif instrument_type == "demat_taxpnl":
        structured = parse_taxpnl(file_bytes, filename)
    elif instrument_type == "demat_tradebook":
        structured = parse_tradebook(file_bytes, filename)
    else:
        raise ValueError(f"Unknown demat instrument type: {instrument_type}")

    journal_entries = generate_journal_entries_rules(structured, instrument_type)
    return structured, journal_entries
