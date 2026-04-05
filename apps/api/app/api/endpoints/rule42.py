"""
Rule 42 ITC Reversal — CRUD + Compute + Annual True-Up API

Persists monthly Rule 42 computations per client, supports
auto-save/load, annual true-up calculation, and Excel export.
"""
from typing import Any, Optional
from uuid import UUID
from datetime import datetime
from dataclasses import asdict

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import desc

from app.api import deps
from app.models.models import User, Rule42Computation, Voucher, VoucherEntry, Ledger
from app.services.gst.rule42_calculator import Rule42Input, calculate_rule42, calculate_rule42_annual_trueup, Rule42Result

import logging
import io
import re

logger = logging.getLogger(__name__)

router = APIRouter()

# GST ledger name patterns for ITC extraction
GST_ITC_PATTERNS = re.compile(
    r'(input\s+cgst|input\s+sgst|input\s+igst|cgst\s+input|sgst\s+input|igst\s+input'
    r'|cgst\s+a/?c|sgst\s+a/?c|igst\s+a/?c'
    r'|central\s+tax|state\s+tax|integrated\s+tax'
    r'|input\s+tax\s+credit)', re.IGNORECASE
)

EXEMPT_PATTERNS = re.compile(
    r'(exempt|nil\s+rated|non[\-\s]?taxable|interest\s+received|interest\s+on'
    r'|rent\s+received|agricultural|exports?\s+without\s+payment)', re.IGNORECASE
)


# ═══════════════════════════════════════════════════════
#  SYNC FROM TALLY — Extract ITC + turnover from vouchers
# ═══════════════════════════════════════════════════════

@router.post("/sync-from-tally")
async def sync_from_tally(
    body: dict,
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_active_user),
) -> Any:
    """
    Extract Rule 42 input data from Tally Purchase & Sales vouchers.

    Parses voucher entries to find actual GST Input Tax Credit amounts
    (CGST/SGST/IGST ledgers) instead of estimating at flat 18%.

    Body: { client_id, company_name, financial_year }
    Returns: { months: [{period, T, T1, T2, T3, E, N, F, source_detail}, ...] }
    """
    client_id = body.get("client_id")
    company_name = body.get("company_name")
    financial_year = body.get("financial_year", "2025-26")

    if not company_name:
        raise HTTPException(status_code=400, detail="company_name is required")

    # Parse FY to date range (e.g. "2025-26" → Apr 2025 to Mar 2026)
    fy_parts = financial_year.split("-")
    fy_start = int(fy_parts[0])
    fy_end = fy_start + 1 if len(fy_parts[1]) == 2 else int(fy_parts[1])
    date_from = f"{fy_start}0401"  # April 1
    date_to = f"{fy_end}0331"      # March 31

    # ── Fetch Purchase vouchers + their entries ──
    from sqlalchemy import and_
    purch_q = select(Voucher).where(
        and_(
            Voucher.company_name == company_name,
            Voucher.voucher_type == "Purchase",
            Voucher.date >= date_from,
            Voucher.date <= date_to,
        )
    ).order_by(Voucher.date)
    purch_result = await db.execute(purch_q)
    purchases = purch_result.scalars().all()

    # Fetch all voucher entries for purchase vouchers
    purch_guids = [v.guid for v in purchases]
    purch_entries_map = {}
    if purch_guids:
        # Process in chunks to avoid too-large IN clause
        for i in range(0, len(purch_guids), 100):
            chunk = purch_guids[i:i+100]
            pe_q = select(VoucherEntry).where(VoucherEntry.voucher_guid.in_(chunk))
            pe_result = await db.execute(pe_q)
            for entry in pe_result.scalars().all():
                purch_entries_map.setdefault(entry.voucher_guid, []).append(entry)

    # ── Fetch Sales vouchers ──
    sales_q = select(Voucher).where(
        and_(
            Voucher.company_name == company_name,
            Voucher.voucher_type == "Sales",
            Voucher.date >= date_from,
            Voucher.date <= date_to,
        )
    ).order_by(Voucher.date)
    sales_result = await db.execute(sales_q)
    sales = sales_result.scalars().all()

    # Fetch sales entries for exempt classification
    sales_guids = [v.guid for v in sales]
    sales_entries_map = {}
    if sales_guids:
        for i in range(0, len(sales_guids), 100):
            chunk = sales_guids[i:i+100]
            se_q = select(VoucherEntry).where(VoucherEntry.voucher_guid.in_(chunk))
            se_result = await db.execute(se_q)
            for entry in se_result.scalars().all():
                sales_entries_map.setdefault(entry.voucher_guid, []).append(entry)

    # ── Build monthly data ──
    months_map = {}

    # Process purchases → extract ITC
    for v in purchases:
        if not v.date:
            continue
        period = _date_to_period(v.date)
        if period not in months_map:
            months_map[period] = {
                "T": 0, "T1": 0, "T2": 0, "T3": 0,
                "E": 0, "N": 0, "F": 0,
                "purchase_count": 0, "sales_count": 0,
                "itc_detail": {"cgst": 0, "sgst": 0, "igst": 0},
            }

        m = months_map[period]
        m["purchase_count"] += 1
        entries = purch_entries_map.get(v.guid, [])

        # Extract GST ITC from ledger entries
        for entry in entries:
            ledger = entry.ledger_name or ""
            amount = abs(float(entry.amount)) if entry.amount else 0

            if GST_ITC_PATTERNS.search(ledger):
                m["T"] += amount
                # Classify by tax type
                if re.search(r'cgst|central', ledger, re.IGNORECASE):
                    m["itc_detail"]["cgst"] += amount
                elif re.search(r'sgst|state', ledger, re.IGNORECASE):
                    m["itc_detail"]["sgst"] += amount
                elif re.search(r'igst|integrated', ledger, re.IGNORECASE):
                    m["itc_detail"]["igst"] += amount

        # If no GST entries found in voucher entries, estimate from amount
        if not any(GST_ITC_PATTERNS.search(e.ledger_name or "") for e in entries):
            estimated_gst = abs(float(v.amount or 0)) * 0.18 / 1.18
            m["T"] += round(estimated_gst, 2)

    # Process sales → classify taxable vs exempt
    for v in sales:
        if not v.date:
            continue
        period = _date_to_period(v.date)
        if period not in months_map:
            months_map[period] = {
                "T": 0, "T1": 0, "T2": 0, "T3": 0,
                "E": 0, "N": 0, "F": 0,
                "purchase_count": 0, "sales_count": 0,
                "itc_detail": {"cgst": 0, "sgst": 0, "igst": 0},
            }

        m = months_map[period]
        m["sales_count"] += 1
        base_amount = abs(float(v.amount or 0))
        entries = sales_entries_map.get(v.guid, [])

        # Check if any entry is GST (→ taxable) or exempt
        has_gst = any(
            re.search(r'cgst|sgst|igst|gst|tax', e.ledger_name or '', re.IGNORECASE)
            for e in entries if not e.is_debit
        )
        is_exempt = EXEMPT_PATTERNS.search(v.party_name or "") or EXEMPT_PATTERNS.search(v.narration or "")

        if is_exempt or not has_gst:
            m["E"] += base_amount  # Exempt turnover
        else:
            # Taxable — subtract GST from total to get base
            gst_in_sale = sum(
                abs(float(e.amount or 0))
                for e in entries
                if re.search(r'cgst|sgst|igst', e.ledger_name or '', re.IGNORECASE)
            )
            m["F"] += base_amount  # Total turnover includes everything
            # We'll compute base_taxable at the end

    # Finalize monthly data
    months_result = []
    for period in sorted(months_map.keys()):
        m = months_map[period]
        # F = Total Turnover (taxable already in F + exempt)
        m["F"] = m.get("F", 0) + m["E"] + m["N"]

        month_label = _period_to_label(period)
        months_result.append({
            "period": period,
            "month": month_label,
            "T": round(m["T"], 2),
            "T1": round(m["T1"], 2),
            "T2": round(m["T2"], 2),
            "T3": round(m["T3"], 2),
            "E": round(m["E"], 2),
            "N": round(m["N"], 2),
            "F": round(m["F"], 2),
            "itc_detail": m["itc_detail"],
            "purchase_count": m["purchase_count"],
            "sales_count": m["sales_count"],
            "source": "tally",
            "auto_filled_fields": ["T", "E", "F"],
        })

    # ── Fallback: If vouchers produced no data, extract from ledger balances ──
    if not months_result:
        logger.info(f"Rule42: No voucher data for {company_name}, trying ledger balances")

        # Query ledgers under Duties & Taxes (Input CGST/SGST/IGST)
        from sqlalchemy import or_
        tax_q = select(Ledger).where(
            Ledger.company_name == company_name,
            or_(
                Ledger.parent.ilike('%Duties%'),
                Ledger.parent.ilike('%Tax%'),
                Ledger.name.ilike('%Input CGST%'),
                Ledger.name.ilike('%Input SGST%'),
                Ledger.name.ilike('%Input IGST%'),
                Ledger.name.ilike('%Input Tax%'),
            )
        )
        tax_result = await db.execute(tax_q)
        tax_ledgers = tax_result.scalars().all()

        total_itc = 0
        itc_detail = {"cgst": 0, "sgst": 0, "igst": 0}
        for led in tax_ledgers:
            name = (led.name or "").lower()
            bal = abs(float(led.closing_balance or 0))
            if any(k in name for k in ['input cgst', 'cgst input', 'central tax input']):
                itc_detail["cgst"] += bal
                total_itc += bal
            elif any(k in name for k in ['input sgst', 'sgst input', 'state tax input']):
                itc_detail["sgst"] += bal
                total_itc += bal
            elif any(k in name for k in ['input igst', 'igst input', 'integrated tax input']):
                itc_detail["igst"] += bal
                total_itc += bal

        # Query sales accounts for turnover
        sales_q = select(Ledger).where(
            Ledger.company_name == company_name,
            or_(
                Ledger.parent.ilike('%Sales%'),
                Ledger.parent.ilike('%Revenue%'),
                Ledger.parent.ilike('%Income%'),
            )
        )
        sales_result = await db.execute(sales_q)
        sales_ledgers = sales_result.scalars().all()

        total_turnover = sum(abs(float(l.closing_balance or 0)) for l in sales_ledgers)

        if total_itc > 0 or total_turnover > 0:
            # Distribute annual figures across 12 months
            monthly_itc = round(total_itc / 12, 2)
            monthly_to = round(total_turnover / 12, 2)

            MONTHS_IN_FY = ['04','05','06','07','08','09','10','11','12','01','02','03']
            for mm in MONTHS_IN_FY:
                yr = fy_start if int(mm) >= 4 else fy_end
                period = f"{yr}-{mm}"
                months_result.append({
                    "period": period,
                    "month": _period_to_label(period),
                    "T": monthly_itc,
                    "T1": 0, "T2": 0, "T3": 0,
                    "E": 0, "N": 0,
                    "F": monthly_to,
                    "itc_detail": {k: round(v/12, 2) for k, v in itc_detail.items()},
                    "purchase_count": 0, "sales_count": 0,
                    "source": "ledger",
                    "auto_filled_fields": ["T", "F"],
                })

            logger.info(f"Rule42 ledger fallback: ITC=₹{total_itc:,.0f}, TO=₹{total_turnover:,.0f}")

    # ── Auto-save to DB if client_id provided ──
    saved_count = 0
    if client_id:
        for md in months_result:
            inp = Rule42Input(
                T=md["T"], T1=md["T1"], T2=md["T2"], T3=md["T3"],
                E=md["E"], N=md["N"], F=md["F"],
                period=md["period"], tax_head="cgst",
            )
            result = calculate_rule42(inp)
            result_dict = asdict(result)

            # Upsert
            existing = await db.execute(
                select(Rule42Computation).where(
                    Rule42Computation.client_id == client_id,
                    Rule42Computation.period == md["period"],
                    Rule42Computation.tax_head == "cgst",
                )
            )
            comp = existing.scalars().first()

            if comp:
                comp.inputs = {"T": md["T"], "T1": md["T1"], "T2": md["T2"], "T3": md["T3"], "E": md["E"], "N": md["N"], "F": md["F"]}
                comp.results = result_dict
                comp.auto_filled_fields = md["auto_filled_fields"]
                comp.status = "draft"
                comp.updated_at = datetime.utcnow()
            else:
                comp = Rule42Computation(
                    client_id=client_id,
                    user_id=current_user.id,
                    firm_id=current_user.firm_id,
                    period=md["period"],
                    financial_year=financial_year,
                    tax_head="cgst",
                    inputs={"T": md["T"], "T1": md["T1"], "T2": md["T2"], "T3": md["T3"], "E": md["E"], "N": md["N"], "F": md["F"]},
                    results=result_dict,
                    status="draft",
                    auto_filled_fields=md["auto_filled_fields"],
                )
                db.add(comp)
            saved_count += 1

            # Add results to response
            md["results"] = result_dict

        await db.commit()

    logger.info(f"Rule42 Tally sync: company={company_name}, months={len(months_result)}, saved={saved_count}")

    return JSONResponse(content={
        "months": months_result,
        "total_months": len(months_result),
        "saved": saved_count,
        "company_name": company_name,
        "financial_year": financial_year,
    })


def _date_to_period(date_str: str) -> str:
    """Convert Tally date '20250415' → '2025-04'"""
    if len(date_str) >= 6:
        return f"{date_str[:4]}-{date_str[4:6]}"
    return date_str


def _period_to_label(period: str) -> str:
    """Convert '2025-04' → 'April 2025'"""
    MONTHS = {
        '01': 'January', '02': 'February', '03': 'March',
        '04': 'April', '05': 'May', '06': 'June',
        '07': 'July', '08': 'August', '09': 'September',
        '10': 'October', '11': 'November', '12': 'December',
    }
    parts = period.split('-')
    return f"{MONTHS.get(parts[1], parts[1])} {parts[0]}" if len(parts) == 2 else period


# ═══════════════════════════════════════════════════════
#  SAVE (Upsert) — Create or Update a computation
# ═══════════════════════════════════════════════════════

@router.post("/save")
async def save_rule42_computation(
    body: dict,
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_active_user),
) -> Any:
    """Save (upsert) a Rule 42 computation for a client+period+tax_head.
    If one already exists, it's updated. Otherwise, a new one is created.

    Body:
        client_id, period, financial_year, tax_head,
        inputs: {T, T1, T2, T3, E, N, F},
        notes (optional), status (optional: draft/final)
    """
    client_id = body.get("client_id")
    period = body.get("period")
    financial_year = body.get("financial_year")
    tax_head = body.get("tax_head", "cgst")
    inputs = body.get("inputs", {})
    notes = body.get("notes")
    status = body.get("status", "draft")
    auto_filled = body.get("auto_filled_fields", [])

    if not client_id or not period or not financial_year:
        raise HTTPException(status_code=400, detail="client_id, period, and financial_year are required")

    # Compute results server-side
    inp = Rule42Input(
        T=float(inputs.get("T", 0)),
        T1=float(inputs.get("T1", 0)),
        T2=float(inputs.get("T2", 0)),
        T3=float(inputs.get("T3", 0)),
        E=float(inputs.get("E", 0)),
        N=float(inputs.get("N", 0)),
        F=float(inputs.get("F", 0)),
        period=period,
        tax_head=tax_head,
    )
    result = calculate_rule42(inp)
    result_dict = asdict(result)

    # Upsert: check if computation exists
    existing = await db.execute(
        select(Rule42Computation).where(
            Rule42Computation.client_id == client_id,
            Rule42Computation.period == period,
            Rule42Computation.tax_head == tax_head,
        )
    )
    comp = existing.scalars().first()

    if comp:
        comp.inputs = inputs
        comp.results = result_dict
        comp.notes = notes
        comp.status = status
        comp.auto_filled_fields = auto_filled
        comp.updated_at = datetime.utcnow()
    else:
        comp = Rule42Computation(
            client_id=client_id,
            user_id=current_user.id,
            firm_id=current_user.firm_id,
            period=period,
            financial_year=financial_year,
            tax_head=tax_head,
            inputs=inputs,
            results=result_dict,
            status=status,
            notes=notes,
            auto_filled_fields=auto_filled,
        )
        db.add(comp)

    await db.commit()
    await db.refresh(comp)

    logger.info(f"Rule42 saved: client={client_id}, period={period}, tax_head={tax_head}, "
                f"reversal=₹{result.total_reversal:,.2f}, eligible=₹{result.net_eligible_itc:,.2f}")

    return JSONResponse(content={
        "id": str(comp.id),
        "period": comp.period,
        "financial_year": comp.financial_year,
        "tax_head": comp.tax_head,
        "inputs": comp.inputs,
        "results": comp.results,
        "status": comp.status,
        "notes": comp.notes,
        "created_at": comp.created_at.isoformat() if comp.created_at else None,
        "updated_at": comp.updated_at.isoformat() if comp.updated_at else None,
    })


# ═══════════════════════════════════════════════════════
#  GET — Load a specific computation
# ═══════════════════════════════════════════════════════

@router.get("/load")
async def load_rule42_computation(
    client_id: str = Query(...),
    period: str = Query(...),
    tax_head: str = Query("cgst"),
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_active_user),
) -> Any:
    """Load a saved computation for a specific client+period+tax_head."""
    result = await db.execute(
        select(Rule42Computation).where(
            Rule42Computation.client_id == client_id,
            Rule42Computation.period == period,
            Rule42Computation.tax_head == tax_head,
        )
    )
    comp = result.scalars().first()

    if not comp:
        return JSONResponse(content={"found": False}, status_code=200)

    return JSONResponse(content={
        "found": True,
        "id": str(comp.id),
        "period": comp.period,
        "financial_year": comp.financial_year,
        "tax_head": comp.tax_head,
        "inputs": comp.inputs,
        "results": comp.results,
        "status": comp.status,
        "notes": comp.notes,
        "auto_filled_fields": comp.auto_filled_fields or [],
        "created_at": comp.created_at.isoformat() if comp.created_at else None,
        "updated_at": comp.updated_at.isoformat() if comp.updated_at else None,
    })


# ═══════════════════════════════════════════════════════
#  HISTORY — All computations for a client
# ═══════════════════════════════════════════════════════

@router.get("/history")
async def list_rule42_history(
    client_id: str = Query(...),
    financial_year: Optional[str] = Query(None),
    tax_head: Optional[str] = Query(None),
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_active_user),
) -> Any:
    """List all saved Rule 42 computations for a client, optionally filtered by FY and tax head."""
    q = select(Rule42Computation).where(
        Rule42Computation.client_id == client_id
    ).order_by(desc(Rule42Computation.period))

    if financial_year:
        q = q.where(Rule42Computation.financial_year == financial_year)
    if tax_head:
        q = q.where(Rule42Computation.tax_head == tax_head)

    result = await db.execute(q)
    comps = result.scalars().all()

    return JSONResponse(content=[
        {
            "id": str(c.id),
            "period": c.period,
            "financial_year": c.financial_year,
            "tax_head": c.tax_head,
            "status": c.status,
            "inputs": c.inputs,
            "results": c.results,
            "notes": c.notes,
            "created_at": c.created_at.isoformat() if c.created_at else None,
            "updated_at": c.updated_at.isoformat() if c.updated_at else None,
        }
        for c in comps
    ])


# ═══════════════════════════════════════════════════════
#  DELETE
# ═══════════════════════════════════════════════════════

@router.delete("/{computation_id}")
async def delete_rule42_computation(
    computation_id: str,
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_active_user),
) -> Any:
    """Delete a saved computation."""
    result = await db.execute(
        select(Rule42Computation).where(Rule42Computation.id == computation_id)
    )
    comp = result.scalars().first()
    if not comp:
        raise HTTPException(status_code=404, detail="Computation not found")

    await db.delete(comp)
    await db.commit()
    return {"ok": True}


# ═══════════════════════════════════════════════════════
#  FINALIZE — Lock a period
# ═══════════════════════════════════════════════════════

@router.post("/{computation_id}/finalize")
async def finalize_rule42(
    computation_id: str,
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_active_user),
) -> Any:
    """Mark a computation as 'final' (locked from casual edits)."""
    result = await db.execute(
        select(Rule42Computation).where(Rule42Computation.id == computation_id)
    )
    comp = result.scalars().first()
    if not comp:
        raise HTTPException(status_code=404, detail="Computation not found")

    comp.status = "final"
    comp.updated_at = datetime.utcnow()
    await db.commit()

    return {"ok": True, "status": "final"}


# ═══════════════════════════════════════════════════════
#  ANNUAL TRUE-UP — Rule 42(2)
# ═══════════════════════════════════════════════════════

@router.get("/annual-trueup")
async def get_annual_trueup(
    client_id: str = Query(...),
    financial_year: str = Query(...),
    tax_head: str = Query("cgst"),
    annual_E: Optional[float] = Query(None),
    annual_N: Optional[float] = Query(None),
    annual_F: Optional[float] = Query(None),
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_active_user),
) -> Any:
    """
    Annual true-up computation under Rule 42(2).
    Aggregates all monthly computations for the FY, then optionally
    recomputes C3 using actual annual turnover figures.
    """
    result = await db.execute(
        select(Rule42Computation).where(
            Rule42Computation.client_id == client_id,
            Rule42Computation.financial_year == financial_year,
            Rule42Computation.tax_head == tax_head,
        ).order_by(Rule42Computation.period)
    )
    comps = result.scalars().all()

    if not comps:
        return JSONResponse(content={"found": False, "message": "No computations found for this FY"})

    # Build monthly summary
    monthly_data = []
    total_T = 0
    total_C2 = 0
    total_C3 = 0
    total_C4 = 0
    total_reversal = 0
    total_E = 0
    total_N = 0
    total_F = 0

    for c in comps:
        r = c.results or {}
        inp = c.inputs or {}
        monthly_data.append({
            "period": c.period,
            "status": c.status,
            "T": float(inp.get("T", 0)),
            "E": float(inp.get("E", 0)),
            "N": float(inp.get("N", 0)),
            "F": float(inp.get("F", 0)),
            "C2": float(r.get("C2", 0)),
            "C3": float(r.get("C3", 0)),
            "C4": float(r.get("C4", 0)),
            "total_reversal": float(r.get("total_reversal", 0)),
            "exempt_ratio": float(r.get("exempt_ratio", 0)),
        })
        total_T += float(inp.get("T", 0))
        total_C2 += float(r.get("C2", 0))
        total_C3 += float(r.get("C3", 0))
        total_C4 += float(r.get("C4", 0))
        total_reversal += float(r.get("total_reversal", 0))
        total_E += float(inp.get("E", 0))
        total_N += float(inp.get("N", 0))
        total_F += float(inp.get("F", 0))

    # If annual actuals are provided, compute true-up
    trueup = None
    if annual_E is not None and annual_F is not None:
        actual_N = annual_N if annual_N is not None else total_N
        F_safe = annual_F if annual_F > 0 else 1.0
        annual_ratio = (annual_E + actual_N) / F_safe
        annual_C3 = total_C2 * annual_ratio
        adjustment = annual_C3 - total_C3

        trueup = {
            "annual_E": round(annual_E, 2),
            "annual_N": round(actual_N, 2),
            "annual_F": round(annual_F, 2),
            "annual_ratio": round(annual_ratio, 6),
            "annual_C3_should_be": round(annual_C3, 2),
            "sum_monthly_C3": round(total_C3, 2),
            "adjustment": round(adjustment, 2),
            "direction": "additional_reversal" if adjustment > 0 else "credit_reclaim",
            "file_in": "April GSTR-3B of next financial year",
        }

    return JSONResponse(content={
        "found": True,
        "financial_year": financial_year,
        "tax_head": tax_head,
        "months_computed": len(comps),
        "monthly_data": monthly_data,
        "totals": {
            "T": round(total_T, 2),
            "E": round(total_E, 2),
            "N": round(total_N, 2),
            "F": round(total_F, 2),
            "C2": round(total_C2, 2),
            "C3": round(total_C3, 2),
            "C4": round(total_C4, 2),
            "total_reversal": round(total_reversal, 2),
        },
        "trueup": trueup,
    })


# ═══════════════════════════════════════════════════════
#  EXCEL EXPORT — Professional Workpaper
# ═══════════════════════════════════════════════════════

@router.get("/export-excel")
async def export_rule42_excel(
    client_id: str = Query(...),
    financial_year: str = Query(...),
    tax_head: str = Query("cgst"),
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_active_user),
) -> Any:
    """Export all Rule 42 computations for a FY as a professional Excel workpaper."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    result = await db.execute(
        select(Rule42Computation).where(
            Rule42Computation.client_id == client_id,
            Rule42Computation.financial_year == financial_year,
            Rule42Computation.tax_head == tax_head,
        ).order_by(Rule42Computation.period)
    )
    comps = result.scalars().all()

    if not comps:
        raise HTTPException(status_code=404, detail="No computations found")

    # Get client name
    from app.models.models import Client
    client_result = await db.execute(select(Client).where(Client.id == client_id))
    client = client_result.scalars().first()
    client_name = client.name if client else "Client"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Rule 42 - {tax_head.upper()}"

    # Styles
    header_font = Font(name='Calibri', size=14, bold=True, color='1E3A8A')
    sub_font = Font(name='Calibri', size=10, color='64748B')
    col_header_font = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
    col_header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    num_font = Font(name='Calibri', size=10)
    total_font = Font(name='Calibri', size=10, bold=True)
    total_fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )

    # Title
    ws.merge_cells('A1:L1')
    ws['A1'] = f"Rule 42 CGST — ITC Reversal Workpaper"
    ws['A1'].font = header_font

    ws.merge_cells('A2:L2')
    ws['A2'] = f"{client_name} | FY {financial_year} | {tax_head.upper()}"
    ws['A2'].font = sub_font

    ws.merge_cells('A3:L3')
    ws['A3'] = f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')} | Rule 42 CGST Rules, 2017"
    ws['A3'].font = Font(name='Calibri', size=8, italic=True, color='94A3B8')

    # Column headers
    headers = ['Period', 'T (Total ITC)', 'T1 (Non-Biz)', 'T2 (Exempt)', 'T3 (Blocked)',
               'E (Exempt TO)', 'N (Non-taxable)', 'F (Total TO)',
               'C2 (Common)', 'Ratio %', 'C3 (Reversal)', 'C4 (Eligible)', 'Status']
    row = 5
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = 14

    ws.column_dimensions['A'].width = 12

    # Data rows
    totals = {k: 0 for k in ['T', 'T1', 'T2', 'T3', 'E', 'N', 'F', 'C2', 'C3', 'C4']}

    for comp in comps:
        row += 1
        inp = comp.inputs or {}
        res = comp.results or {}

        vals = [
            comp.period,
            float(inp.get('T', 0)), float(inp.get('T1', 0)),
            float(inp.get('T2', 0)), float(inp.get('T3', 0)),
            float(inp.get('E', 0)), float(inp.get('N', 0)), float(inp.get('F', 0)),
            float(res.get('C2', 0)),
            round(float(res.get('exempt_ratio', 0)) * 100, 2),
            float(res.get('C3', 0)), float(res.get('C4', 0)),
            comp.status.upper(),
        ]

        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = num_font
            cell.border = thin_border
            if isinstance(v, float):
                cell.number_format = '#,##0'
            if col == 10:  # ratio
                cell.number_format = '0.00"%"'

        for k, ci in [('T', 2), ('T1', 3), ('T2', 4), ('T3', 5), ('E', 6), ('N', 7), ('F', 8),
                       ('C2', 9), ('C3', 11), ('C4', 12)]:
            totals[k] += float(inp.get(k, 0)) if k in ['T', 'T1', 'T2', 'T3', 'E', 'N', 'F'] else float(res.get(k, 0))

    # Totals row
    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = total_font
    for k, ci in [('T', 2), ('T1', 3), ('T2', 4), ('T3', 5), ('E', 6), ('N', 7), ('F', 8),
                   ('C2', 9), ('C3', 11), ('C4', 12)]:
        cell = ws.cell(row=row, column=ci, value=round(totals[k], 2))
        cell.font = total_font
        cell.fill = total_fill
        cell.number_format = '#,##0'
        cell.border = thin_border

    # Annual ratio
    if totals['F'] > 0:
        annual_ratio = (totals['E'] + totals['N']) / totals['F']
        cell = ws.cell(row=row, column=10, value=round(annual_ratio * 100, 2))
        cell.font = total_font
        cell.fill = total_fill

    # Notes section
    row += 2
    ws.cell(row=row, column=1, value="Notes:").font = Font(name='Calibri', size=9, bold=True)
    row += 1
    ws.cell(row=row, column=1, value="• Monthly C3 is provisional. Annual true-up required under Rule 42(2).").font = Font(name='Calibri', size=8, color='64748B')
    row += 1
    ws.cell(row=row, column=1, value="• Adjustment to be filed in April GSTR-3B of next FY or via annual return.").font = Font(name='Calibri', size=8, color='64748B')
    row += 1
    ws.cell(row=row, column=1, value="• References: GSTR-3B Table 4(B)(1) for D1, Table 4(B)(2) for D2+C3.").font = Font(name='Calibri', size=8, color='64748B')

    # Write to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Rule42_ITC_{client_name.replace(' ', '_')}_{financial_year}_{tax_head.upper()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""},
    )
