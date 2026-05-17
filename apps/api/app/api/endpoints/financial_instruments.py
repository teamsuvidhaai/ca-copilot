"""
Financial Instruments API
─────────────────────────
Endpoints for uploading and parsing Demat, Mutual Fund, and PMS statements.
AI extracts holdings, transactions, dividends, capital gains → journal entries.
"""

import json
import logging
import uuid
from datetime import datetime
from decimal import Decimal
from typing import Any, Optional

from app.services.fi_rule_parsers import (
    parse_demat_markdown, parse_pms_markdown, parse_26as_markdown,
    generate_journal_entries_for_demat, generate_journal_entries_for_pms,
)
from app.services.cas_parser import parse_cas_markdown, generate_journal_entries_from_parsed
from fastapi import APIRouter, BackgroundTasks, Depends, File, Form, HTTPException, UploadFile
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.api import deps
from app.core.config import settings
from app.models.models import User

router = APIRouter()
logger = logging.getLogger(__name__)

MAX_FILE_SIZE = 25 * 1024 * 1024  # 25 MB


# ─── Claude Structuring Prompts ────────────────────────

DEMAT_PROMPT = """You are a financial data extraction AI specialising in Indian Demat account statements (CDSL/NSDL).

Given the raw text from a Demat statement, extract ALL holdings and transactions into strict JSON.

**Rules:**
1. Extract EVERY holding and transaction — do not skip any row.
2. Dates must be ISO format: "YYYY-MM-DD"
3. Amounts must be plain numbers (no commas, no ₹) or null.
4. Classify each transaction: Buy, Sell, Bonus, Split, Rights, Transfer In, Transfer Out, IPO Allotment, Dividend
5. For capital gains, compute buy_value, sell_value, gain_loss where possible.

Return this exact JSON:
{
    "dp_id": "string or null",
    "client_id": "string or null",
    "depository": "CDSL or NSDL or null",
    "statement_date": "YYYY-MM-DD or null",
    "holdings": [
        {
            "isin": "string",
            "scrip_name": "string",
            "quantity": number,
            "avg_cost": number or null,
            "market_value": number or null,
            "market_price": number or null
        }
    ],
    "transactions": [
        {
            "date": "YYYY-MM-DD",
            "isin": "string or null",
            "scrip_name": "string",
            "type": "Buy/Sell/Bonus/Dividend/etc",
            "quantity": number or null,
            "price": number or null,
            "amount": number or null,
            "buy_value": number or null,
            "sell_value": number or null,
            "gain_loss": number or null,
            "holding_period": "Short Term or Long Term or null"
        }
    ],
    "dividends": [
        {
            "date": "YYYY-MM-DD",
            "scrip_name": "string",
            "amount": number,
            "tds_deducted": number or null
        }
    ],
    "capital_gains_summary": {
        "short_term_gain": number or null,
        "long_term_gain": number or null,
        "total_gain": number or null
    }
}"""

MF_PROMPT = """You are a financial data extraction AI specialising in Indian Mutual Fund statements (CAS / AMC statements).

Given the raw text from a Mutual Fund statement, extract ALL fund holdings and transactions.

**Rules:**
1. Extract EVERY fund, transaction, and dividend — do not skip any.
2. Dates: "YYYY-MM-DD". Amounts: plain numbers, no commas/₹.
3. Classify transactions: Purchase, Redemption, SIP, STP, Switch In, Switch Out, Dividend Payout, Dividend Reinvestment
4. Compute capital gains where possible.

Return this exact JSON:
{
    "investor_name": "string or null",
    "pan": "string or null",
    "statement_period_start": "YYYY-MM-DD or null",
    "statement_period_end": "YYYY-MM-DD or null",
    "funds": [
        {
            "fund_name": "string",
            "amc": "string or null",
            "folio_number": "string or null",
            "scheme_type": "Equity/Debt/Hybrid/ELSS/Liquid/etc",
            "units": number or null,
            "nav": number or null,
            "market_value": number or null,
            "cost_value": number or null
        }
    ],
    "transactions": [
        {
            "date": "YYYY-MM-DD",
            "fund_name": "string",
            "type": "Purchase/Redemption/SIP/STP/Switch/Dividend",
            "units": number or null,
            "nav": number or null,
            "amount": number or null,
            "stamp_duty": number or null,
            "stt": number or null
        }
    ],
    "dividends": [
        {
            "date": "YYYY-MM-DD",
            "fund_name": "string",
            "amount": number,
            "tds_deducted": number or null
        }
    ],
    "capital_gains_summary": {
        "short_term_gain": number or null,
        "long_term_gain": number or null,
        "total_gain": number or null
    }
}"""

PMS_PROMPT = """You are a financial data extraction AI specialising in Indian Portfolio Management Service (PMS) statements.

Given the raw text from a PMS statement, extract ALL holdings, transactions, fees, and performance data.

**Rules:**
1. Extract EVERY holding, transaction, fee entry — do not skip any.
2. Dates: "YYYY-MM-DD". Amounts: plain numbers only.
3. Classify transactions: Buy, Sell, Dividend, Corporate Action, Fee Deduction
4. Track PMS-specific fields: management fees, performance fees, brokerage.

Return this exact JSON:
{
    "portfolio_name": "string or null",
    "pms_provider": "string or null",
    "client_name": "string or null",
    "statement_date": "YYYY-MM-DD or null",
    "portfolio_value": number or null,
    "invested_value": number or null,
    "holdings": [
        {
            "scrip_name": "string",
            "isin": "string or null",
            "quantity": number,
            "avg_cost": number or null,
            "market_price": number or null,
            "market_value": number or null,
            "weight_pct": number or null,
            "unrealised_gain": number or null
        }
    ],
    "transactions": [
        {
            "date": "YYYY-MM-DD",
            "scrip_name": "string",
            "type": "Buy/Sell/Dividend/Corporate Action/Fee",
            "quantity": number or null,
            "price": number or null,
            "amount": number or null,
            "brokerage": number or null,
            "gain_loss": number or null
        }
    ],
    "fees": [
        {
            "date": "YYYY-MM-DD",
            "type": "Management Fee/Performance Fee/Brokerage/Custodian Fee/Other",
            "amount": number,
            "gst_on_fee": number or null
        }
    ],
    "capital_gains_summary": {
        "short_term_gain": number or null,
        "long_term_gain": number or null,
        "total_gain": number or null
    }
}"""

JOURNAL_ENTRY_PROMPT = """You are an expert Indian Chartered Accountant. Given the structured financial instrument data below,
generate proper double-entry journal entries following Indian accounting standards.

**Rules:**
1. Use standard Indian ledger names (e.g., "Investment in Equity Shares", "Dividend Income", "STCG on Shares", "LTCG on Equity MF")
2. Each entry must balance (total Dr = total Cr)
3. Include narration for each entry
4. Handle TDS on dividends (TDS Receivable Dr, Dividend Income Cr)
5. For capital gains, separate STCG and LTCG
6. For PMS fees, debit "PMS Management Fees" / "PMS Performance Fees" and credit the PMS account
7. Assign the correct Tally voucher_type:
   - Share/MF purchases → "Purchase"
   - Share/MF sales → "Sales"
   - Dividend/interest received → "Receipt"
   - Brokerage/DP charges/fees paid → "Payment"
   - Capital gain/loss entries, transfers, adjustments → "Journal"

Return JSON:
{
    "journal_entries": [
        {
            "date": "YYYY-MM-DD",
            "voucher_type": "Purchase" or "Sales" or "Receipt" or "Payment" or "Journal",
            "narration": "string",
            "ledger_entries": [
                {"ledger_name": "string", "amount": number, "side": "Dr" or "Cr"}
            ]
        }
    ]
}"""


# ─── Helper: Extract text from uploaded file ───────────
async def _extract_text(file_bytes: bytes, filename: str) -> str:
    """Extract text from PDF/Excel using LlamaParse."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "csv":
        try:
            return file_bytes.decode("utf-8", errors="replace")
        except Exception:
            return file_bytes.decode("latin-1", errors="replace")

    if ext in ("xlsx", "xls"):
        # Parse Excel properly using openpyxl
        import io
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            all_text = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    rows.append(",".join(cells))
                all_text.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows))
            wb.close()
            result = "\n\n".join(all_text)
            logger.info(f"Excel extraction: {len(result)} chars from {len(wb.sheetnames)} sheets")
            return result
        except Exception as e:
            logger.error(f"openpyxl failed: {e}, falling back to raw decode")
            return file_bytes.decode("latin-1", errors="replace")

    # PDF — use LlamaParse
    if not settings.LLAMA_CLOUD_API_KEY:
        raise ValueError("LLAMA_CLOUD_API_KEY not configured")

    try:
        from llama_cloud import AsyncLlamaCloud
    except ImportError:
        raise ImportError("llama_cloud package not installed. Run: pip install llama_cloud>=1.0")

    import tempfile
    client = AsyncLlamaCloud(api_key=settings.LLAMA_CLOUD_API_KEY)

    with tempfile.NamedTemporaryFile(suffix=f"_{filename}", delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name

    logger.info(f"Uploading {filename} ({len(file_bytes)} bytes) to LlamaParse...")

    try:
        file_obj = await client.files.create(file=tmp_path, purpose="parse")
        result = await client.parsing.parse(
            file_id=file_obj.id, tier="agentic", version="latest",
            output_options={"markdown": {"tables": {"output_tables_as_markdown": True}}},
            processing_options={"ocr_parameters": {"languages": ["en"]}},
            expand=["text", "markdown"],
        )

        pages_text = []
        if result.markdown and result.markdown.pages:
            pages_text = [p.markdown or "" for p in result.markdown.pages]
        elif result.text and result.text.pages:
            pages_text = [p.text or "" for p in result.text.pages]

        full_text = "\n\n--- PAGE BREAK ---\n\n".join(pages_text)
        logger.info(f"LlamaParse: {len(full_text)} chars from {len(pages_text)} pages")
        return full_text
    except Exception as e:
        logger.error(f"LlamaParse extraction failed: {e}")
        raise


def _structure_with_rules(text: str, instrument_type: str) -> dict:
    """Structure extracted text using rule-based parsers (no AI cost)."""
    if instrument_type == "demat" or instrument_type.startswith("demat_"):
        return parse_demat_markdown(text)
    elif instrument_type == "mutual_fund":
        return parse_cas_markdown(text)
    elif instrument_type == "pms" or instrument_type.startswith("pms_"):
        return parse_pms_markdown(text)
    else:
        # Default to demat parser for unknown types
        return parse_demat_markdown(text)


def _merge_structured_results(results: list, instrument_type: str) -> dict:
    """Merge structured data from multiple chunks into one."""
    merged = {}
    list_keys = {"holdings", "transactions", "dividends", "fees", "funds"}

    for r in results:
        if not r:
            continue
        for key, value in r.items():
            if key in list_keys and isinstance(value, list):
                merged.setdefault(key, []).extend(value)
            elif key == "capital_gains_summary" and isinstance(value, dict):
                existing = merged.get("capital_gains_summary", {})
                for k, v in value.items():
                    if v is not None:
                        existing[k] = (existing.get(k) or 0) + (v or 0)
                merged["capital_gains_summary"] = existing
            elif key not in merged and value is not None:
                merged[key] = value  # keep first non-null scalar (e.g., dp_id, client_name)

    return merged


def _generate_journal_entries_rules(structured_data: dict, instrument_type: str) -> dict:
    """Generate journal entries using rule-based accounting logic (no AI cost)."""
    if instrument_type == "mutual_fund":
        entries = generate_journal_entries_from_parsed(structured_data)
    elif instrument_type == "pms" or instrument_type.startswith("pms_"):
        entries = generate_journal_entries_for_pms(structured_data)
    else:
        entries = generate_journal_entries_for_demat(structured_data)

    validated = _validate_journal_entries(entries)
    return {"journal_entries": validated}



def _validate_journal_entries(entries: list) -> list:
    """Validate AI-generated journal entries: balance check, amount sanity, date format."""
    valid_voucher_types = {"Purchase", "Sales", "Receipt", "Payment", "Journal"}
    validated = []
    for je in entries:
        ledger_entries = je.get("ledger_entries", [])
        if not ledger_entries:
            continue

        # Clean up amounts — force numeric
        for le in ledger_entries:
            try:
                le["amount"] = round(float(le.get("amount", 0)), 2)
            except (ValueError, TypeError):
                le["amount"] = 0
            if le.get("side") not in ("Dr", "Cr"):
                le["side"] = "Dr"  # default

        # Remove zero-amount entries
        ledger_entries = [le for le in ledger_entries if abs(le["amount"]) > 0.01]
        if len(ledger_entries) < 2:
            continue  # need at least one Dr and one Cr

        je["ledger_entries"] = ledger_entries

        # Balance check: total Dr must equal total Cr
        total_dr = sum(le["amount"] for le in ledger_entries if le["side"] == "Dr")
        total_cr = sum(le["amount"] for le in ledger_entries if le["side"] == "Cr")
        if abs(total_dr - total_cr) > 1.0:  # Allow ₹1 rounding tolerance
            je["narration"] = f"⚠ UNBALANCED (Dr={total_dr:.2f}, Cr={total_cr:.2f}) — " + (je.get("narration") or "")
            logger.warning(f"Unbalanced journal entry: Dr={total_dr}, Cr={total_cr}")

        # Validate voucher_type
        if je.get("voucher_type") not in valid_voucher_types:
            je["voucher_type"] = "Journal"  # safe default

        validated.append(je)

    return validated


# ─── Upload PDF to Supabase ────────────────────────────
def _upload_to_supabase(file_bytes: bytes, upload_id: str, filename: str) -> str:
    try:
        from supabase import create_client
        client = create_client(settings.SUPABASE_URL, settings.SUPABASE_KEY)
        bucket = "financial-instruments"
        try:
            client.storage.create_bucket(bucket, options={"public": False})
        except Exception:
            pass
        storage_path = f"{upload_id}/{filename}"
        content_type = "application/pdf" if filename.lower().endswith(".pdf") else "application/octet-stream"
        client.storage.from_(bucket).upload(
            path=storage_path, file=file_bytes,
            file_options={"content-type": content_type, "upsert": "true"}
        )
        return storage_path
    except Exception as e:
        logger.warning(f"Supabase upload failed (non-fatal): {e}")
        return ""


# ─── DB-backed storage ─────────────────────────────────
# Uses FinancialInstrumentUpload model — survives restarts, supports queries

from app.db.session import AsyncSessionLocal
from app.models.models import FinancialInstrumentUpload, FIEntry
import hashlib


# ─── Background processor ─────────────────────────────
async def _process_upload(upload_id: str, file_bytes: bytes, filename: str, instrument_type: str):
    """Background: extract → structure → journal entries. Writes progress to DB."""
    async with AsyncSessionLocal() as db:
        try:
            row = (await db.execute(select(FinancialInstrumentUpload).where(FinancialInstrumentUpload.id == upload_id))).scalar_one()
            row.status = "extracting"
            await db.commit()

            _upload_to_supabase(file_bytes, upload_id, filename)

            ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
            is_demat_excel = instrument_type.startswith("demat_") and ext in ("xlsx", "xls")

            if is_demat_excel:
                # ═══ DIRECT PARSING — No AI needed ═══
                from app.api.endpoints.demat_parser import parse_demat_excel
                row.status = "structuring"
                await db.commit()

                structured, entries = parse_demat_excel(file_bytes, filename, instrument_type)
                row.structured_data = structured
                row.raw_text = f"[Direct parsed — {len(structured.get('transactions', structured.get('holdings', [])))} items]"
                row.status = "generating_entries"
                await db.commit()

                row.journal_entries = entries
                row.journal_entry_count = len(entries)
                
                # Save to fi_entries table
                for je in entries:
                    fi_entry = FIEntry(
                        upload_id=upload_id,
                        client_id=row.client_id,
                        date=datetime.strptime(je.get("date"), "%Y-%m-%d").date() if je.get("date") else None,
                        narration=je.get("narration"),
                        scrip=je.get("scrip"),
                        trade_count=je.get("tradeCount", 0),
                        cg_type=je.get("cgType"),
                        voucher_type=je.get("voucher_type", "Journal"),
                        status=je.get("status", "draft"),
                        total_amount=je.get("total_amount"),
                        entries=je.get("entries", je.get("ledger_entries", []))
                    )
                    db.add(fi_entry)

                row.status = "completed"
                await db.commit()

                logger.info(f"✅ FI Upload {upload_id} (direct): {len(entries)} journal entries")
            else:
                # ═══ AI PIPELINE — for PDFs and other formats ═══
                raw_text = await _extract_text(file_bytes, filename)
                row.status = "structuring"
                row.raw_text = raw_text[:2000]
                await db.commit()

                base_type = instrument_type.split('_')[0] if instrument_type.startswith('demat_') else instrument_type

                # Rule-based parser (no AI cost)
                structured = _structure_with_rules(raw_text, base_type)
                row.structured_data = structured
                row.status = "generating_entries"
                await db.commit()

                journal = _generate_journal_entries_rules(structured, base_type)
                entries = journal.get("journal_entries", [])

                row.journal_entries = entries
                row.journal_entry_count = len(entries)
                
                # Save to fi_entries table
                for je in entries:
                    fi_entry = FIEntry(
                        upload_id=upload_id,
                        client_id=row.client_id,
                        date=datetime.strptime(je.get("date"), "%Y-%m-%d").date() if je.get("date") else None,
                        narration=je.get("narration"),
                        scrip=je.get("scrip"),
                        trade_count=je.get("tradeCount", 0),
                        cg_type=je.get("cgType"),
                        voucher_type=je.get("voucher_type", "Journal"),
                        status=je.get("status", "draft"),
                        total_amount=je.get("total_amount"),
                        entries=je.get("entries", je.get("ledger_entries", []))
                    )
                    db.add(fi_entry)

                row.status = "completed"
                await db.commit()

                logger.info(f"✅ FI Upload {upload_id}: {len(entries)} journal entries")

        except Exception as e:
            logger.error(f"❌ FI Upload {upload_id} failed: {e}")
            row = (await db.execute(select(FinancialInstrumentUpload).where(FinancialInstrumentUpload.id == upload_id))).scalar_one_or_none()
            if row:
                row.status = "failed"
                row.error_message = str(e)[:2000]
                await db.commit()


# ─── Upload Endpoint ──────────────────────────────────
@router.post("/upload")
async def upload_statement(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    client_id: str = Form(...),
    instrument_type: str = Form(...),  # demat, mutual_fund, pms
    current_user: User = Depends(deps.get_current_user),
    db: AsyncSession = Depends(deps.get_db),
) -> Any:
    """Upload a Demat/MF/PMS statement for AI processing."""

    valid_types = ("demat", "mutual_fund", "pms", "demat_holdings", "demat_taxpnl", "demat_tradebook")
    if instrument_type not in valid_types:
        raise HTTPException(400, f"instrument_type must be one of: {', '.join(valid_types)}")

    if not file.filename:
        raise HTTPException(400, "No file provided")

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("pdf", "xlsx", "xls", "csv"):
        raise HTTPException(400, f"Supported formats: PDF, XLSX, XLS, CSV. Got .{ext}")

    file_bytes = await file.read()
    if len(file_bytes) > MAX_FILE_SIZE:
        raise HTTPException(400, f"File too large ({len(file_bytes)/(1024*1024):.1f} MB). Max 25 MB.")
    if len(file_bytes) == 0:
        raise HTTPException(400, "File is empty")

    # Compute file hash for duplicate detection
    file_hash = hashlib.sha256(file_bytes).hexdigest()

    # Check for duplicate: same file already uploaded for this client
    existing = (await db.execute(
        select(FinancialInstrumentUpload).where(
            FinancialInstrumentUpload.client_id == client_id,
            FinancialInstrumentUpload.file_hash == file_hash,
        )
    )).scalar_one_or_none()
    if existing:
        raise HTTPException(
            409,
            f"Duplicate file — this statement was already uploaded as \"{existing.filename}\" "
            f"on {existing.created_at.strftime('%d %b %Y') if existing.created_at else 'unknown date'}."
        )

    upload_id = str(uuid.uuid4())
    row = FinancialInstrumentUpload(
        id=upload_id,
        client_id=client_id,
        user_id=str(current_user.id),
        instrument_type=instrument_type,
        filename=file.filename,
        file_hash=file_hash,
        status="processing",
    )
    db.add(row)
    await db.commit()

    background_tasks.add_task(_process_upload, upload_id, file_bytes, file.filename, instrument_type)

    return {
        "id": upload_id,
        "status": "processing",
        "message": f"Processing {file.filename} as {instrument_type}. This may take 30-60 seconds.",
    }


# ─── Status Endpoint ──────────────────────────────────
@router.get("/status/{upload_id}")
async def get_upload_status(
    upload_id: str,
    current_user: User = Depends(deps.get_current_user),
    db: AsyncSession = Depends(deps.get_db),
) -> Any:
    """Check the status of a financial instrument upload."""
    row = (await db.execute(
        select(FinancialInstrumentUpload).where(FinancialInstrumentUpload.id == upload_id)
    )).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "Upload not found")
    return {
        "id": str(row.id),
        "status": row.status,
        "instrument_type": row.instrument_type,
        "filename": row.filename,
        "error": row.error_message,
        "created_at": row.created_at.isoformat() if row.created_at else None,
        "has_data": row.structured_data is not None,
        "journal_entry_count": row.journal_entry_count or 0,
        "je_status": row.je_status or "pending",
    }


# ─── Get Structured Data ─────────────────────────────
@router.get("/data/{upload_id}")
async def get_structured_data(
    upload_id: str,
    current_user: User = Depends(deps.get_current_user),
    db: AsyncSession = Depends(deps.get_db),
) -> Any:
    """Get the AI-parsed structured data from an upload."""
    row = (await db.execute(
        select(FinancialInstrumentUpload).where(FinancialInstrumentUpload.id == upload_id)
    )).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "Upload not found")
    if row.status not in ("completed", "generating_entries"):
        raise HTTPException(400, f"Data not ready. Current status: {row.status}")
    return {
        "instrument_type": row.instrument_type,
        "structured_data": row.structured_data,
    }


# ─── Get Journal Entries ──────────────────────────────
@router.get("/journal-entries/{upload_id}")
async def get_journal_entries(
    upload_id: str,
    current_user: User = Depends(deps.get_current_user),
    db: AsyncSession = Depends(deps.get_db),
) -> Any:
    """Get generated journal entries for an upload."""
    row = (await db.execute(
        select(FinancialInstrumentUpload).where(FinancialInstrumentUpload.id == upload_id)
    )).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "Upload not found")
    if row.status != "completed":
        raise HTTPException(400, f"Entries not ready. Status: {row.status}")

    # Try fetching from the new fi_entries table first
    stmt = select(FIEntry).where(FIEntry.upload_id == upload_id)
    result = await db.execute(stmt)
    db_entries = result.scalars().all()
    
    if db_entries:
        entries = []
        for e in db_entries:
            entries.append({
                "id": str(e.id),
                "date": e.date.isoformat() if e.date else None,
                "narration": e.narration,
                "scrip": e.scrip,
                "tradeCount": e.trade_count,
                "cgType": e.cg_type,
                "voucher_type": e.voucher_type,
                "status": e.status,
                "total_amount": float(e.total_amount) if e.total_amount else 0,
                "entries": e.entries
            })
        return {"journal_entries": entries}

    # Return stored entries from JSONB column if available
    if row.journal_entries:
        return {"journal_entries": row.journal_entries}

    # Fallback: dynamically generate from structured_data (e.g. PMS uploads)
    if row.structured_data:
        result = _generate_journal_entries_rules(row.structured_data, row.instrument_type)
        entries = result.get("journal_entries", [])
        if entries:
            # Cache for next time
            row.journal_entries = entries
            row.journal_entry_count = len(entries)
            await db.commit()
            return {"journal_entries": entries}

    return {"journal_entries": []}


# ─── List Uploads ─────────────────────────────────────
@router.get("/")
async def list_uploads(
    client_id: str = None,
    instrument_type: str = None,
    date_from: str = None,
    date_to: str = None,
    current_user: User = Depends(deps.get_current_user),
    db: AsyncSession = Depends(deps.get_db),
) -> Any:
    """List all financial instrument uploads.
    
    Optional date_from/date_to in YYYYMMDD format filter by created_at.
    This lets the frontend scope uploads to a selected financial year.
    """
    query = select(FinancialInstrumentUpload)
    if client_id:
        query = query.where(FinancialInstrumentUpload.client_id == client_id)
    if instrument_type:
        query = query.where(FinancialInstrumentUpload.instrument_type == instrument_type)

    # Date range filtering (YYYYMMDD → datetime)
    if date_from:
        try:
            dt_from = datetime.strptime(date_from, "%Y%m%d")
            query = query.where(FinancialInstrumentUpload.created_at >= dt_from)
        except ValueError:
            pass
    if date_to:
        try:
            from datetime import timedelta
            dt_to = datetime.strptime(date_to, "%Y%m%d") + timedelta(days=1)
            query = query.where(FinancialInstrumentUpload.created_at < dt_to)
        except ValueError:
            pass

    query = query.order_by(FinancialInstrumentUpload.created_at.desc())

    result = await db.execute(query)
    rows = result.scalars().all()
    return [
        {
            "id": str(r.id),
            "client_id": str(r.client_id),
            "instrument_type": r.instrument_type,
            "filename": r.filename,
            "status": r.status,
            "created_at": r.created_at.isoformat() if r.created_at else None,
            "journal_entry_count": r.journal_entry_count or 0,
            "je_status": r.je_status or "pending",
            "pms_account_id": str(r.pms_account_id) if r.pms_account_id else None,
        }
        for r in rows
    ]


# ─── Delete Upload (with Supabase cleanup) ────────────
@router.delete("/{upload_id}")
async def delete_upload(
    upload_id: str,
    current_user: User = Depends(deps.get_current_user),
    db: AsyncSession = Depends(deps.get_db),
) -> Any:
    """Delete a financial instrument upload and its Supabase file."""
    row = (await db.execute(
        select(FinancialInstrumentUpload).where(FinancialInstrumentUpload.id == upload_id)
    )).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "Upload not found")

    # Clean up Supabase storage
    storage_path = f"{upload_id}/{row.filename}"
    try:
        from supabase import create_client
        client = create_client(settings.SUPABASE_URL, settings.SUPABASE_KEY)
        client.storage.from_("financial-instruments").remove([storage_path])
        logger.info(f"🗑️ Deleted Supabase file: {storage_path}")
    except Exception as e:
        logger.warning(f"Supabase cleanup failed (non-fatal): {e}")

    await db.delete(row)
    await db.commit()
    return {"message": "Deleted"}


# ─── Get PDF URL ──────────────────────────────────
@router.get("/pdf/{upload_id}")
async def get_upload_pdf(
    upload_id: str,
    current_user: User = Depends(deps.get_current_user),
    db: AsyncSession = Depends(deps.get_db),
) -> Any:
    """Get a signed URL for the uploaded statement PDF."""
    row = (await db.execute(
        select(FinancialInstrumentUpload).where(FinancialInstrumentUpload.id == upload_id)
    )).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "Upload not found")

    storage_path = f"{upload_id}/{row.filename}"
    try:
        from supabase import create_client
        client = create_client(settings.SUPABASE_URL, settings.SUPABASE_KEY)
        signed = client.storage.from_("financial-instruments").create_signed_url(
            path=storage_path, expires_in=3600
        )
        url = signed.get("signedURL") or signed.get("signedUrl") or ""
        if not url:
            raise ValueError("No signed URL returned")
        return {"url": url, "filename": row.filename}
    except Exception as e:
        logger.warning(f"PDF URL generation failed: {e}")
        raise HTTPException(400, f"Could not generate PDF URL: {str(e)}")


# ─── Approve / Sync Journal Entries ───────────────────
from pydantic import BaseModel as BaseModel  # noqa: E402 (re-import for scope)

class JeStatusUpdate(BaseModel):
    je_status: str  # pending, approved, synced


@router.patch("/{upload_id}/je-status")
async def update_je_status(
    upload_id: str,
    payload: JeStatusUpdate,
    current_user: User = Depends(deps.get_current_user),
    db: AsyncSession = Depends(deps.get_db),
) -> Any:
    """Update the journal entry approval status of an upload (pending → approved → synced)."""
    if payload.je_status not in ("pending", "approved", "synced"):
        raise HTTPException(400, "je_status must be pending, approved, or synced")

    row = (await db.execute(
        select(FinancialInstrumentUpload).where(FinancialInstrumentUpload.id == upload_id)
    )).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "Upload not found")

    row.je_status = payload.je_status
    await db.commit()
    logger.info(f"Upload {upload_id} je_status -> {payload.je_status}")
    return {"id": upload_id, "je_status": payload.je_status, "message": f"Journal entries marked as {payload.je_status}"}


class EntryStatusUpdate(BaseModel):
    status: str  # draft, approved, synced


@router.patch("/journal-entry/{entry_id}/status")
async def update_entry_status(
    entry_id: str,
    payload: EntryStatusUpdate,
    current_user: User = Depends(deps.get_current_user),
    db: AsyncSession = Depends(deps.get_db),
) -> Any:
    """Update the status of a single journal entry in fi_entries."""
    if payload.status not in ("draft", "approved", "synced"):
        raise HTTPException(400, "status must be draft, approved, or synced")

    row = (await db.execute(
        select(FIEntry).where(FIEntry.id == entry_id)
    )).scalar_one_or_none()
    
    if not row:
        raise HTTPException(404, "Entry not found")

    row.status = payload.status
    await db.commit()
    logger.info(f"FIEntry {entry_id} status -> {payload.status}")
    return {"id": entry_id, "status": payload.status, "message": f"Entry marked as {payload.status}"}


# ═══════════════════════════════════════════════════════
# MANUAL JOURNAL ENTRIES — CRUD for hand-keyed FI vouchers
# Uses fi_uploads table with instrument_type='manual_entry'
# ═══════════════════════════════════════════════════════

from pydantic import BaseModel
from typing import List

class ManualEntryLedger(BaseModel):
    ledger_name: str
    group: str = ""
    amount: float
    side: str  # "Dr" or "Cr"

class ManualEntryCreate(BaseModel):
    client_id: str
    txn_type: str  # purchase, sale, dividend, charges, sip, ipo, bonus, transfer
    date: str  # YYYY-MM-DD
    scrip: str
    narration: str = ""
    voucher_type: str = "Journal"
    status: str = "draft"  # draft, approved, synced
    total_amount: float = 0
    entries: List[ManualEntryLedger]

class ManualEntryStatusUpdate(BaseModel):
    status: str  # draft, approved, synced


@router.post("/manual-entry")
async def create_manual_entry(
    payload: ManualEntryCreate,
    current_user: User = Depends(deps.get_current_user),
    db: AsyncSession = Depends(deps.get_db),
) -> Any:
    """Create a manual FI journal entry. Persisted in fi_uploads as instrument_type='manual_entry'."""
    if not payload.scrip:
        raise HTTPException(400, "Scrip / Fund name is required")
    if len(payload.entries) < 2:
        raise HTTPException(400, "Need at least 2 ledger entries (Dr + Cr)")

    upload_id = str(uuid.uuid4())
    journal_entry = {
        "date": payload.date,
        "scrip": payload.scrip,
        "txn_type": payload.txn_type,
        "voucher_type": payload.voucher_type,
        "narration": payload.narration,
        "status": payload.status,
        "total_amount": payload.total_amount,
        "entries": [e.dict() for e in payload.entries],
        "created_at": datetime.utcnow().isoformat(),
    }

    row = FinancialInstrumentUpload(
        id=upload_id,
        client_id=payload.client_id,
        user_id=str(current_user.id),
        instrument_type="manual_entry",
        filename=f"Manual — {payload.scrip}",
        file_hash=hashlib.sha256(f"{payload.client_id}-{payload.scrip}-{payload.date}-{upload_id}".encode()).hexdigest(),
        status="completed",
        structured_data={"txn_type": payload.txn_type, "scrip": payload.scrip},
        journal_entries=[journal_entry],
        journal_entry_count=1,
    )
    db.add(row)
    await db.commit()

    logger.info(f"✅ Manual FI entry created: {upload_id} — {payload.scrip} ({payload.txn_type})")
    return {
        "id": upload_id,
        "status": payload.status,
        "scrip": payload.scrip,
        "txn_type": payload.txn_type,
        "message": f"Manual entry saved — {payload.scrip}",
    }


@router.get("/manual-entries")
async def list_manual_entries(
    client_id: str = None,
    status: str = None,
    current_user: User = Depends(deps.get_current_user),
    db: AsyncSession = Depends(deps.get_db),
) -> Any:
    """List all manual FI journal entries for a client."""
    query = select(FinancialInstrumentUpload).where(
        FinancialInstrumentUpload.instrument_type == "manual_entry"
    )
    if client_id:
        query = query.where(FinancialInstrumentUpload.client_id == client_id)
    query = query.order_by(FinancialInstrumentUpload.created_at.desc())

    result = await db.execute(query)
    rows = result.scalars().all()

    entries = []
    for r in rows:
        je = (r.journal_entries or [{}])[0] if r.journal_entries else {}
        entry_status = je.get("status", "draft")
        if status and entry_status != status:
            continue
        entries.append({
            "id": str(r.id),
            "client_id": str(r.client_id),
            "txn_type": je.get("txn_type", ""),
            "scrip": je.get("scrip", ""),
            "date": je.get("date", ""),
            "narration": je.get("narration", ""),
            "voucher_type": je.get("voucher_type", "Journal"),
            "status": entry_status,
            "total_amount": je.get("total_amount", 0),
            "entries": je.get("entries", []),
            "created_at": r.created_at.isoformat() if r.created_at else None,
        })
    return entries


@router.patch("/manual-entry/{entry_id}/status")
async def update_manual_entry_status(
    entry_id: str,
    payload: ManualEntryStatusUpdate,
    current_user: User = Depends(deps.get_current_user),
    db: AsyncSession = Depends(deps.get_db),
) -> Any:
    """Update the status of a manual journal entry (draft → approved → synced)."""
    if payload.status not in ("draft", "approved", "synced"):
        raise HTTPException(400, "Status must be draft, approved, or synced")

    row = (await db.execute(
        select(FinancialInstrumentUpload).where(
            FinancialInstrumentUpload.id == entry_id,
            FinancialInstrumentUpload.instrument_type == "manual_entry",
        )
    )).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "Manual entry not found")

    # Update status inside JSONB
    entries = row.journal_entries or [{}]
    if entries:
        entries[0]["status"] = payload.status
    row.journal_entries = entries
    # Force SQLAlchemy to detect the JSONB change
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(row, "journal_entries")
    await db.commit()

    return {"id": entry_id, "status": payload.status, "message": f"Status updated to {payload.status}"}


@router.delete("/manual-entry/{entry_id}")
async def delete_manual_entry(
    entry_id: str,
    current_user: User = Depends(deps.get_current_user),
    db: AsyncSession = Depends(deps.get_db),
) -> Any:
    """Delete a manual FI journal entry."""
    row = (await db.execute(
        select(FinancialInstrumentUpload).where(
            FinancialInstrumentUpload.id == entry_id,
            FinancialInstrumentUpload.instrument_type == "manual_entry",
        )
    )).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "Manual entry not found")

    await db.delete(row)
    await db.commit()
    return {"message": "Manual entry deleted"}


# ═══════════════════════════════════════════════════════
# 26AS / AIS — Upload, Extract, Auto-Match
# ═══════════════════════════════════════════════════════

TDS_26AS_PROMPT = """You are an expert Indian tax data extraction AI specialising in Form 26AS and AIS (Annual Information Statement).

Given the raw text from a 26AS or AIS document, extract ALL TDS entries organised by section.

**Rules:**
1. Extract EVERY entry from Part A (TDS) and Part A1 (TDS 15G/15H). Do NOT skip any.
2. Dates must be ISO format: "YYYY-MM-DD"
3. Amounts must be plain numbers (no commas, no ₹) or null.
4. Group entries by TDS section code (194, 194A, 194K, 194DA, 194B, etc.)
5. Also extract Part B (TCS), Part C (Tax Paid), and SFT entries if present.

Return this exact JSON:
{
    "pan": "string or null",
    "assessment_year": "string or null",
    "financial_year": "string or null",
    "tds_entries": [
        {
            "section": "194 or 194A or 194K or 194DA or 194B etc",
            "section_description": "Dividend / Interest / MF Income etc",
            "tan_of_deductor": "string or null",
            "deductor_name": "string",
            "transaction_date": "YYYY-MM-DD or null",
            "amount_paid_credited": number,
            "tds_deducted": number,
            "tds_deposited": number or null
        }
    ],
    "tcs_entries": [
        {
            "section": "string",
            "collector_name": "string",
            "amount": number,
            "tcs_collected": number
        }
    ],
    "tax_paid": [
        {
            "type": "Advance Tax or Self Assessment or TDS",
            "bsr_code": "string or null",
            "date": "YYYY-MM-DD",
            "amount": number,
            "challan_serial": "string or null"
        }
    ],
    "sft_entries": [
        {
            "transaction_type": "Purchase of shares / Sale of shares / MF Purchase / MF Redemption etc",
            "reported_by": "string",
            "amount": number,
            "count_of_transactions": number or null
        }
    ],
    "summary": {
        "total_tds": number,
        "total_tcs": number,
        "total_tax_paid": number,
        "total_income_reported": number
    }
}"""


async def _process_26as_upload(upload_id: str, file_bytes: bytes, filename: str, client_id: str):
    """Background: extract 26AS → structure → auto-match with existing statements."""
    async with AsyncSessionLocal() as db:
        try:
            row = (await db.execute(
                select(FinancialInstrumentUpload).where(FinancialInstrumentUpload.id == upload_id)
            )).scalar_one()
            row.status = "extracting"
            await db.commit()

            _upload_to_supabase(file_bytes, upload_id, filename)

            raw_text = await _extract_text(file_bytes, filename)
            row.status = "structuring"
            row.raw_text = raw_text[:2000]
            await db.commit()

            # Rule-based 26AS extraction (no AI cost)
            structured = parse_26as_markdown(raw_text)

            # Run auto-matching
            row.status = "generating_entries"
            await db.commit()

            match_results = await _match_26as_with_statements(db, structured, client_id)
            structured["match_results"] = match_results

            row.structured_data = structured
            row.journal_entries = []
            row.journal_entry_count = 0
            row.status = "completed"
            await db.commit()

            logger.info(f"✅ 26AS Upload {upload_id}: {len(structured.get('tds_entries', []))} TDS entries, "
                        f"{match_results.get('matched_count', 0)} matched, "
                        f"{match_results.get('unmatched_count', 0)} unmatched")

        except Exception as e:
            logger.error(f"❌ 26AS Upload {upload_id} failed: {e}")
            row = (await db.execute(
                select(FinancialInstrumentUpload).where(FinancialInstrumentUpload.id == upload_id)
            )).scalar_one_or_none()
            if row:
                row.status = "failed"
                row.error_message = str(e)[:2000]
                await db.commit()


async def _match_26as_with_statements(db: AsyncSession, data_26as: dict, client_id: str) -> dict:
    """Match 26AS TDS entries against uploaded Demat/MF/PMS statements."""

    tds_entries = data_26as.get("tds_entries", [])
    if not tds_entries:
        return {"matched": [], "unmatched_26as": [], "unmatched_stmts": [],
                "mismatched": [], "matched_count": 0, "unmatched_count": 0,
                "mismatch_count": 0, "section_summary": {}}

    # Fetch all completed FI uploads for this client (non-26AS)
    stmt_rows = (await db.execute(
        select(FinancialInstrumentUpload).where(
            FinancialInstrumentUpload.client_id == client_id,
            FinancialInstrumentUpload.status == "completed",
            FinancialInstrumentUpload.instrument_type != "26as",
        )
    )).scalars().all()

    # Collect all dividends & TDS from statements
    stmt_dividends = []  # {"source": "demat/mf/pms", "name": "...", "amount": X, "tds": Y}
    stmt_cg = {"stcg": 0, "ltcg": 0}

    for sr in stmt_rows:
        sd = sr.structured_data or {}
        utype = sr.instrument_type

        # Dividends
        for div in sd.get("dividends", []):
            amt = float(div.get("amount", 0) or 0)
            tds = float(div.get("tds_deducted", div.get("tds", 0)) or 0)
            name = div.get("scrip_name", div.get("fund_name", div.get("security_name", "Unknown")))
            if amt > 0:
                stmt_dividends.append({
                    "source": utype, "name": name, "amount": amt,
                    "tds": tds, "date": div.get("date", div.get("ex_date", ""))
                })

        # Capital gains summary
        cg = sd.get("capital_gains_summary", {})
        stmt_cg["stcg"] += float(cg.get("short_term_gain", 0) or 0)
        stmt_cg["ltcg"] += float(cg.get("long_term_gain", 0) or 0)

    # Group 26AS entries by section
    section_groups = {}
    for entry in tds_entries:
        sec = entry.get("section", "unknown")
        section_groups.setdefault(sec, []).append(entry)

    matched = []
    unmatched_26as = []
    mismatched = []

    # Process each section
    remaining_stmt_divs = list(stmt_dividends)

    for entry in tds_entries:
        sec = entry.get("section", "")
        deductor = entry.get("deductor_name", "")
        amt_26as = float(entry.get("amount_paid_credited", 0) or 0)
        tds_26as = float(entry.get("tds_deducted", 0) or 0)

        # Try to match dividend entries (Sec 194, 194K, 194DA)
        if sec in ("194", "194K", "194DA"):
            best_match = None
            best_idx = -1
            best_diff = float("inf")

            for i, sd in enumerate(remaining_stmt_divs):
                # Match by amount (within 10% tolerance) and fuzzy name
                if sd["amount"] > 0:
                    diff = abs(sd["amount"] - amt_26as)
                    pct_diff = diff / max(amt_26as, 1) * 100

                    if pct_diff < 15 and diff < best_diff:
                        best_match = sd
                        best_idx = i
                        best_diff = diff

            if best_match is not None:
                tds_diff = abs(best_match["tds"] - tds_26as)
                status = "matched" if tds_diff < 10 else "tds_mismatch"

                result_entry = {
                    "section": sec,
                    "deductor": deductor,
                    "amount_26as": amt_26as,
                    "tds_26as": tds_26as,
                    "amount_stmt": best_match["amount"],
                    "tds_stmt": best_match["tds"],
                    "stmt_source": best_match["source"],
                    "stmt_name": best_match["name"],
                    "amount_diff": round(amt_26as - best_match["amount"], 2),
                    "tds_diff": round(tds_26as - best_match["tds"], 2),
                }

                if status == "matched":
                    matched.append(result_entry)
                else:
                    result_entry["issue"] = f"TDS mismatch: 26AS ₹{tds_26as:.0f} vs Stmt ₹{best_match['tds']:.0f}"
                    mismatched.append(result_entry)

                remaining_stmt_divs.pop(best_idx)
            else:
                unmatched_26as.append({
                    "section": sec,
                    "deductor": deductor,
                    "amount_26as": amt_26as,
                    "tds_26as": tds_26as,
                    "issue": "No matching dividend found in uploaded statements"
                })
        else:
            # Other sections (194A interest, etc.) — report as informational
            unmatched_26as.append({
                "section": sec,
                "deductor": deductor,
                "amount_26as": amt_26as,
                "tds_26as": tds_26as,
                "issue": f"Section {sec} — not auto-matched (manual verification needed)"
            })

    # Remaining statement dividends that weren't matched
    unmatched_stmts = [{
        "source": sd["source"], "name": sd["name"],
        "amount": sd["amount"], "tds": sd["tds"],
        "issue": "Dividend in statement but not found in 26AS"
    } for sd in remaining_stmt_divs]

    # Build section summary
    section_summary = {}
    for sec, entries in section_groups.items():
        total_amount = sum(float(e.get("amount_paid_credited", 0) or 0) for e in entries)
        total_tds = sum(float(e.get("tds_deducted", 0) or 0) for e in entries)
        section_summary[sec] = {
            "count": len(entries),
            "total_amount": round(total_amount, 2),
            "total_tds": round(total_tds, 2),
        }

    return {
        "matched": matched,
        "unmatched_26as": unmatched_26as,
        "unmatched_stmts": unmatched_stmts,
        "mismatched": mismatched,
        "matched_count": len(matched),
        "unmatched_count": len(unmatched_26as) + len(unmatched_stmts),
        "mismatch_count": len(mismatched),
        "section_summary": section_summary,
    }


# ─── 26AS Upload Endpoint ────────────────────────────
@router.post("/26as-upload")
async def upload_26as(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    client_id: str = Form(...),
    current_user: User = Depends(deps.get_current_user),
    db: AsyncSession = Depends(deps.get_db),
) -> Any:
    """Upload a 26AS or AIS PDF for TDS reconciliation."""
    if not file.filename:
        raise HTTPException(400, "No file provided")

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("pdf", "xlsx", "xls", "csv"):
        raise HTTPException(400, f"Supported: PDF, Excel, CSV. Got .{ext}")

    file_bytes = await file.read()
    if len(file_bytes) > MAX_FILE_SIZE:
        raise HTTPException(400, f"File too large ({len(file_bytes)/(1024*1024):.1f} MB). Max 25 MB.")
    if len(file_bytes) == 0:
        raise HTTPException(400, "File is empty")

    file_hash = hashlib.sha256(file_bytes).hexdigest()

    # Check duplicate
    existing = (await db.execute(
        select(FinancialInstrumentUpload).where(
            FinancialInstrumentUpload.client_id == client_id,
            FinancialInstrumentUpload.file_hash == file_hash,
        )
    )).scalar_one_or_none()
    if existing:
        raise HTTPException(409, f"Duplicate — already uploaded as \"{existing.filename}\"")

    upload_id = str(uuid.uuid4())
    row = FinancialInstrumentUpload(
        id=upload_id,
        client_id=client_id,
        user_id=str(current_user.id),
        instrument_type="26as",
        filename=file.filename,
        file_hash=file_hash,
        status="processing",
    )
    db.add(row)
    await db.commit()

    background_tasks.add_task(_process_26as_upload, upload_id, file_bytes, file.filename, client_id)

    return {
        "id": upload_id,
        "status": "processing",
        "message": f"Processing 26AS: {file.filename}. Extraction + auto-matching may take 30-60 seconds.",
    }


# ─── 26AS Match Results ──────────────────────────────
@router.get("/26as-match")
async def get_26as_match(
    client_id: str,
    current_user: User = Depends(deps.get_current_user),
    db: AsyncSession = Depends(deps.get_db),
) -> Any:
    """Get auto-match results from the latest 26AS upload for a client."""
    row = (await db.execute(
        select(FinancialInstrumentUpload).where(
            FinancialInstrumentUpload.client_id == client_id,
            FinancialInstrumentUpload.instrument_type == "26as",
        ).order_by(FinancialInstrumentUpload.created_at.desc())
    )).scalars().first()

    if not row:
        return {"status": "not_uploaded", "message": "No 26AS uploaded for this client"}
    if row.status != "completed":
        return {"status": row.status, "message": f"26AS processing: {row.status}"}

    sd = row.structured_data or {}
    return {
        "status": "completed",
        "upload_id": str(row.id),
        "filename": row.filename,
        "uploaded_at": row.created_at.isoformat() if row.created_at else None,
        "pan": sd.get("pan"),
        "assessment_year": sd.get("assessment_year"),
        "financial_year": sd.get("financial_year"),
        "tds_entries_count": len(sd.get("tds_entries", [])),
        "summary": sd.get("summary", {}),
        "match_results": sd.get("match_results", {}),
        "section_summary": sd.get("match_results", {}).get("section_summary", {}),
    }


